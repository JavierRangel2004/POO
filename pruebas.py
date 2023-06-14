import pandas as pd
import datetime
import openpyxl
import random

def excelOutput(schedules):
    wb = openpyxl.Workbook() # Create a blank workbook
    hourrowrelation={
        '07:00': 2,
        '07:30': 3,
        '08:00': 4,
        '08:30': 5,
        '09:00': 6,
        '09:30': 7,
        '10:00': 8,
        '10:30': 9,
        '11:00': 10,
        '11:30': 11,
        '12:00': 12,
        '12:30': 13,
        '13:00': 14,
        '13:30': 15,
        '14:00': 16,
        '14:30': 17,
        '15:00': 18,
        '15:30': 19,
        '16:00': 20,
        '16:30': 21,
        '17:00': 22,
        '17:30': 23,
        '18:00': 24,
        '18:30': 25,
        '19:00': 26,
        '19:30': 27,
        '20:00': 28,
        '20:30': 29,
        '21:00': 30,
        '21:30': 31,
    }
    schedulecount = 1
    for schedule in schedules:
        sheet=wb.create_sheet(title="Horario " + str(schedulecount))
        sheet = wb['Horario ' + str(schedulecount)]
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 30
        #change the height of the rows 2:32 to 30 pixels
        for i in range(2, 32):
            sheet.row_dimensions[i].height = 30
        sheet['A1'] = "Hora"
        sheet['B1'] = 'Lunes'
        sheet['C1'] = 'Martes'
        sheet['D1'] = 'Miércoles'
        sheet['E1'] = 'Jueves'
        sheet['F1'] = 'Viernes'
        # Fill from A2 to A31 with blocks of 30 minutes starting at 7:00 (7:00 - 7:30, 7:30 - 8:00, etc.)
        for i in range(2, 32):
            sheet['A' + str(i)] = f"{6 + i // 2}:{30 * (i % 2):02d} - {6 + (i + 1) // 2}:{30 * ((i + 1) % 2):02d}"
        days = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
        row = 2
        for day in days:
            if day in schedule:
                classes = schedule[day]
                for class_info in classes:
                    for class_period, details in class_info.items():
                        #align the block of time with the corresponding row using the hourrowrelation dictionary
                        start_time, end_time = class_period.split(' - ')
                        start_cell = sheet.cell(row=hourrowrelation[start_time], column=days.index(day) + 2) #
                        end_cell = sheet.cell(row=hourrowrelation[end_time]-1, column=days.index(day) + 2)
                        merge_range = start_cell.coordinate + ':' + end_cell.coordinate
                        sheet.merge_cells(merge_range)
                        #fill the block of time with the color of the class
                        sheet[start_cell.coordinate].fill = openpyxl.styles.PatternFill(start_color=details['Color'], end_color=details['Color'], fill_type='solid')
                        sheet[start_cell.coordinate] = f"{start_time} - {end_time} {details['Materia']} {details['Maestro']} ID Salon: {details['Salon']}"
                        sheet[start_cell.coordinate].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')                       
        #make the cells A1:F31 have a border
        for i in range(1, 32):
            for j in range(1, 7):
                sheet.cell(row=i, column=j).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        #make the cells A2:F31 adjust the text to fit in the cell and center it horizontally and vertically

        for i in range(2, 32):
            for j in range(1, 7):
                sheet.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        #fill the cells A1:F1 with a yellow background
        for i in range(1, 7):
            sheet.cell(row=1, column=i).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        schedulecount += 1
    #delete default sheet
    wb.remove(wb['Sheet'])
    wb.save('Horariosss.xlsx')

schedules =[#lista de horarios
    [#lista de clases - schedule
        {#clase - group
            'id': 20,
            'classNumber': 4448,
            'group': '2540',
            'language': 'Español',
            'students': '30/32',
            'modality': 'Presencial',
            'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.',
            'startDate': None, 
            'endDate': None, 
            'options': 0, 
            'status': True, 
            'creationDate': datetime.date(2023, 6, 14), 
            'lastupDate': '2023-06-14 00:30:14', 
            'subject': 'COMPING COM106 - Procesamiento de Imágenes', 
            'teacher': 'Eduardo Fragoso Navarro', 
            'schedules': 
            [#lista de bloques por clase - block
                {#horario y lugar
                    'id': 19, 
                    'classroomID': 11, 
                    'day': 'Lun', 
                    'startTime': '08:30:00', 
                    'endTime': '10:00:00', 
                    'status': True, 
                    'creationDate': '2023-06-14 00:00:00', 
                    'lastupDate': '2023-06-14 00:30:14'
                },
                {
                    'id': 20, 
                    'classroomID': 11, 
                    'day': 'Jue', 
                    'startTime': '08:30:00', 
                    'endTime': '10:00:00', 
                    'status': True, 
                    'creationDate': '2023-06-14 00:00:00', 
                    'lastupDate': '2023-06-14 00:30:14'
                },
                {
                    'id': 213, 
                    'classroomID': 11, 
                    'day': 'Lun', 
                    'startTime': '08:30:00', 
                    'endTime': '10:00:00', 
                    'status': True, 
                    'creationDate': 
                    '2023-06-14 00:00:00', 
                    'lastupDate': '2023-06-14 01:03:39'
                }, 
                {'id': 214,'classroomID': 11, 'day': 'Jue', 'startTime': '08:30:00', 'endTime': '10:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}
            ]
        }, {'id': 45, 'classNumber': 4025, 'group': '191', 'language': 'Español', 'students': '21/37', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:15', 'subject': 'FORHUM HUM003 - Ética', 'teacher': 'Jorge Uriel Gutiérrez Rojas', 'schedules': [{'id': 66, 'classroomID': 23, 'day': 'Mart', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 67, 'classroomID': 24, 'day': 'V', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 260, 'classroomID': 23, 'day': 'Mart', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:40'}, {'id': 261, 'classroomID': 24, 'day': 'V', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:40'}]}, {'id': 23, 'classNumber': 4451, 'group': '2271', 'language': 'Español', 'students': '26/33', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:15', 'subject': 'COMPING COM112 - Introducción a las Bases de Datos', 'teacher': 'Bárbaro Jorge Ferro Castro', 'schedules': [{'id': 26, 'classroomID': 10, 'day': 'V', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 27, 'classroomID': 7, 'day': 'Mart', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 28, 'classroomID': 7, 'day': 'Lun', 'startTime': '10:30:00', 'endTime': '12:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 220, 'classroomID': 10, 'day': 'Mart', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 221, 'classroomID': 7, 'day': 'Lun', 'startTime': '10:30:00', 'endTime': '12:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 222, 'classroomID': 7, 'day': 'V', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}]}, {'id': 16, 'classNumber': 4443, 'group': '1673', 'language': 'Español', 'students': '13/24', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:14', 'subject': 'COMPING COM104 - Programación y Estructura de Datos', 'teacher': 'Gerardo Bárcena Ruíz', 'schedules': [{'id': 12, 'classroomID': 8, 'day': 'V', 'startTime': '07:00:00', 'endTime': '08:30:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 13, 'classroomID': 10, 'day': 'Jue', 'startTime': '15:00:00', 'endTime': '16:30:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 14, 'classroomID': 1, 'day': 'Miérc', 'startTime': '19:30:00', 'endTime': '21:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 15, 'classroomID': 1, 'day': 'Mart', 'startTime': '07:00:00', 'endTime': '08:30:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 206, 'classroomID': 8, 'day': 'V', 'startTime': '07:00:00', 'endTime': '08:30:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 207, 'classroomID': 10, 'day': 'Jue', 'startTime': '15:00:00', 'endTime': '16:30:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 208, 'classroomID': 1, 'day': 'Miérc', 'startTime': '19:30:00', 'endTime': '21:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 209, 'classroomID': 1, 'day': 'Mart', 'startTime': '07:00:00', 'endTime': '08:30:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}]}], [{'id': 20, 'classNumber': 4448, 'group': '2540', 'language': 'Español', 'students': '30/32', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:14', 'subject': 'COMPING COM106 - Procesamiento de Imágenes', 'teacher': 'Eduardo Fragoso Navarro', 'schedules': [{'id': 19, 'classroomID': 11, 'day': 'Lun', 'startTime': '08:30:00', 'endTime': '10:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 20, 'classroomID': 11, 'day': 'Jue', 'startTime': '08:30:00', 'endTime': '10:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 213, 'classroomID': 11, 'day': 'Lun', 'startTime': '08:30:00', 'endTime': '10:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 214, 'classroomID': 11, 'day': 'Jue', 'startTime': '08:30:00', 'endTime': '10:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}]}, {'id': 45, 'classNumber': 4025, 'group': '191', 'language': 'Español', 'students': '21/37', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:15', 'subject': 'FORHUM HUM003 - Ética', 'teacher': 'Jorge Uriel Gutiérrez Rojas', 'schedules': [{'id': 66, 'classroomID': 23, 'day': 'Mart', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 67, 'classroomID': 24, 'day': 'V', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 260, 'classroomID': 23, 'day': 'Mart', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:40'}, {'id': 261, 'classroomID': 24, 'day': 'V', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:40'}]}, {'id': 23, 'classNumber': 4451, 'group': '2271', 'language': 'Español', 'students': '26/33', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:15', 'subject': 'COMPING COM112 - Introducción a las Bases de Datos', 'teacher': 'Bárbaro Jorge Ferro Castro', 'schedules': [{'id': 26, 'classroomID': 10, 'day': 'V', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 27, 'classroomID': 7, 'day': 'Mart', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 28, 'classroomID': 7, 'day': 'Lun', 'startTime': '10:30:00', 'endTime': '12:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 220, 'classroomID': 10, 'day': 'Mart', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 221, 'classroomID': 7, 'day': 'Lun', 'startTime': '10:30:00', 'endTime': '12:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 222, 'classroomID': 7, 'day': 'V', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}]}, {'id': 17, 'classNumber': 4444, 'group': '3351', 'language': 'Español', 'students': '23/30', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:14', 'subject': 'COMPING COM104 - Programación y Estructura de Datos', 'teacher': 'Gerardo Bárcena Ruíz', 'schedules': [{'id': 16, 'classroomID': 9, 'day': 'Mart', 'startTime': '14:30:00', 'endTime': '16:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 17, 'classroomID': 8, 'day': 'V', 'startTime': '14:30:00', 'endTime': '16:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 18, 'classroomID': 6, 'day': 'Lun', 'startTime': '14:30:00', 'endTime': '16:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 210, 'classroomID': 9, 'day': 'V', 'startTime': '14:30:00', 'endTime': '16:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 211, 'classroomID': 8, 'day': 'Lun', 'startTime': '14:30:00', 'endTime': '16:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 212, 'classroomID': 6, 'day': 'Mart', 'startTime': '14:30:00', 'endTime': '16:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}]}], [{'id': 20, 'classNumber': 4448, 'group': '2540', 'language': 'Español', 'students': '30/32', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:14', 'subject': 'COMPING COM106 - Procesamiento de Imágenes', 'teacher': 'Eduardo Fragoso Navarro', 'schedules': [{'id': 19, 'classroomID': 11, 'day': 'Lun', 'startTime': '08:30:00', 'endTime': '10:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 20, 'classroomID': 11, 'day': 'Jue', 'startTime': '08:30:00', 'endTime': '10:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:14'}, {'id': 213, 'classroomID': 11, 'day': 'Lun', 'startTime': '08:30:00', 'endTime': '10:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 214, 'classroomID': 11, 'day': 'Jue', 'startTime': '08:30:00', 'endTime': '10:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}]}, {'id': 45, 'classNumber': 4025, 'group': '191', 'language': 'Español', 'students': '21/37', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:15', 'subject': 'FORHUM HUM003 - Ética', 'teacher': 'Jorge Uriel Gutiérrez Rojas', 'schedules': [{'id': 66, 'classroomID': 23, 'day': 'Mart', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 67, 'classroomID': 24, 'day': 'V', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 260, 'classroomID': 23, 'day': 'Mart', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:40'}, {'id': 261, 'classroomID': 24, 'day': 'V', 'startTime': '09:30:00', 'endTime': '11:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:40'}]}, {'id': 23, 'classNumber': 4451, 'group': '2271', 'language': 'Español', 'students': '26/33', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:15', 'subject': 'COMPING COM112 - Introducción a las Bases de Datos', 'teacher': 'Bárbaro Jorge Ferro Castro', 'schedules': [{'id': 26, 'classroomID': 10, 'day': 'V', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 27, 'classroomID': 7, 'day': 'Mart', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 28, 'classroomID': 7, 'day': 'Lun', 'startTime': '10:30:00', 'endTime': '12:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:15'}, {'id': 220, 'classroomID': 10, 'day': 'Mart', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 221, 'classroomID': 7, 'day': 'Lun', 'startTime': '10:30:00', 'endTime': '12:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}, {'id': 222, 'classroomID': 7, 'day': 'V', 'startTime': '11:30:00', 'endTime': '13:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:39'}]}, {'id': 140, 'classNumber': 4453, 'group': '789', 'language': 'Español', 'students': '17/24', 'modality': 'Presencial', 'description': 'Se prevé que las sesiones se lleven a cabo en el campus de manera presencial.', 'startDate': None, 'endDate': None, 'options': 0, 'status': True, 'creationDate': datetime.date(2023, 6, 14), 'lastupDate': '2023-06-14 00:30:18', 'subject': 'MATEING MAT106 - Ecuaciones Diferenciales', 'teacher': 'Pavel Real Pérez', 'schedules': [{'id': 182, 'classroomID': 38, 'day': 'Jue', 'startTime': '14:30:00', 'endTime': '16:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:18'}, {'id': 183, 'classroomID': 38, 'day': 'Lun', 'startTime': '14:30:00', 'endTime': '16:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 00:30:18'}, {'id': 376, 'classroomID': 38, 'day': 'Lun', 'startTime': '14:30:00', 'endTime': '16:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:43'}, {'id': 377, 'classroomID': 38, 'day': 'Jue', 'startTime': '14:30:00', 'endTime': '16:00:00', 'status': True, 'creationDate': '2023-06-14 00:00:00', 'lastupDate': '2023-06-14 01:03:43'}]}]]


def cleanSchedulesOutput(schedules):
    cleanSchedules = []
    for schedule in schedules:
        Horario = {
            'Lunes': [],
            'Martes': [],
            'Miercoles': [],
            'Jueves': [],
            'Viernes': []
        }
        colorblockoptions = {
        "Morado": "c778ff",
        "Amarillo": "fffd78",
        "Rojo": "ff8178",
        "Verde": "95ff78",
        "Azul": "78fffa",
        "Naranja": "ffc478",
        "Rosa": "ff78db",
        "Verde Claro": "78ffae",
        "Azul Obscuro": "787fff",
        "Gris":"bab8ba"
        }
        for group in schedule:
            #select a random color for the group without repeating
            color=colorblockoptions.popitem()[1]
            for block in group['schedules']:
                block['startTime'] = block['startTime'][:-3]
                block['endTime'] = block['endTime'][:-3]
                hora=block['startTime']+' - '+block['endTime']
                if block['day']=='Lun':
                    Horario['Lunes'].append({hora:{'Clase': group['classNumber'],'Materia': group['subject'],'Maestro': group['teacher'], 'Salon': block['classroomID'], 'Color': color}})
                elif block['day']=='Mart':
                    Horario['Martes'].append({hora:{'Clase': group['classNumber'],'Materia': group['subject'],'Maestro': group['teacher'], 'Salon': block['classroomID'], 'Color': color}})
                elif block['day']=='Miérc':
                    Horario['Miercoles'].append({hora:{'Clase': group['classNumber'],'Materia': group['subject'],'Maestro': group['teacher'], 'Salon': block['classroomID'], 'Color': color}})
                elif block['day']=='Jue':
                    Horario['Jueves'].append({hora:{'Clase': group['classNumber'],'Materia': group['subject'],'Maestro': group['teacher'], 'Salon': block['classroomID'], 'Color': color}})
                elif block['day']=='V':
                    Horario['Viernes'].append({hora:{'Clase': group['classNumber'],'Materia': group['subject'],'Maestro': group['teacher'], 'Salon': block['classroomID'], 'Color': color}})
        cleanSchedules.append(Horario)
    return cleanSchedules

result = cleanSchedulesOutput(schedules)

excelOutput(result)
